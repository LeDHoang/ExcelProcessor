#!/usr/bin/env python3

import argparse
import hashlib
import json
import logging
import os
import re
import shutil
import sys
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple
import subprocess
import time as _time

# Third-party imports are imported lazily where possible to reduce startup time

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)


@dataclass
class SheetAssetPaths:
    pdf_path: Optional[Path]
    png_path: Optional[Path]
    ocr_path: Optional[Path]


@dataclass
class ExtractedSheet:
    workbook_id: str
    sheet_index: int
    sheet_name: str
    cells: List[Dict]
    named_ranges: List[Dict]
    images: List[Dict]
    smartart_drawings: List[str]
    assets: Dict[str, Optional[str]]
    links_to: List[str]


# -----------------------------
# Filesystem helpers
# -----------------------------

def sanitize_name(name: str) -> str:
    safe = re.sub(r"[^A-Za-z0-9._-]", "_", name)
    # Avoid reserved names and overly long filenames
    return safe[:120] if len(safe) > 120 else safe


def ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


# -----------------------------
# Hash & workbook load
# -----------------------------

def compute_sha256(file_path: Path) -> str:
    logger.info(f"Computing SHA-256 hash for {file_path.name}")
    start_time = time.time()
    h = hashlib.sha256()
    with file_path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    elapsed = time.time() - start_time
    logger.info(f"Hash computed in {elapsed:.2f}s: sha256:{h.hexdigest()[:16]}...")
    return f"sha256:{h.hexdigest()}"


# -----------------------------
# Structured extraction (openpyxl)
# -----------------------------

def load_values_workbook(path: Path):
    from openpyxl import load_workbook

    logger.info(f"Loading workbook (values only): {path.name}")
    start_time = time.time()
    values_wb = load_workbook(
        filename=str(path), read_only=True, data_only=True, keep_links=True
    )
    elapsed = time.time() - start_time
    logger.info(f"Values workbook loaded in {elapsed:.2f}s")
    return values_wb


def a1_from_row_col(row: int, col: int) -> str:
    from openpyxl.utils import get_column_letter

    return f"{get_column_letter(col)}{row}"


def regex_cross_sheet_refs(text: str) -> List[str]:
    # Extract referenced sheet names from formula text
    # Matches: =... Sheet Name!A1 or 'Sheet Name'!A1 or Sheet_1!A1:Z9
    pattern = re.compile(r"(?i)=.*?([A-Za-z0-9_ ]+)!\$?[A-Z]+\$?\d+(?::\$?[A-Z]+\$?\d+)?")
    names: List[str] = []
    for m in pattern.finditer(text or ""):
        names.append(m.group(1).strip().strip("'\""))
    # Deduplicate
    unique = []
    for n in names:
        if n not in unique:
            unique.append(n)
    return unique


def is_effectively_visible_state(state: str) -> bool:
    return state == "visible"


def parse_defined_names_from_workbook(xlsx_path: Path) -> List[Dict]:
    # Parse xl/workbook.xml definedNames for named ranges
    import zipfile
    from lxml import etree

    named_ranges: List[Dict] = []
    try:
        with zipfile.ZipFile(xlsx_path, 'r') as zf:
            if 'xl/workbook.xml' not in zf.namelist():
                return named_ranges
            wb_root = etree.fromstring(zf.read('xl/workbook.xml'))
            ns = {'a': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
            for dn in wb_root.xpath('.//a:definedNames/a:definedName', namespaces=ns):
                name = dn.get('name')
                attr_text = (dn.text or '').strip()
                local_sheet_id = dn.get('localSheetId')
                named_ranges.append({
                    'name': name,
                    'attr_text': attr_text,
                    'localSheetId': local_sheet_id
                })
    except Exception as e:
        logger.warning(f"Failed to parse defined names: {e}")
    return named_ranges


def extract_sheet_structured(values_ws, sheet_name: str, formulas_map: Dict[str, str], hyperlinks_map: Dict[str, str], comments_map: Dict[str, str]) -> Tuple[List[Dict], List[str]]:
    logger.info(f"Extracting structured data from sheet: {sheet_name}")
    start_time = time.time()
    
    cells: List[Dict] = []
    links_to: List[str] = []

    # Iterate using meta_ws for full metadata; pull values from values_ws
    # We cannot rely on UsedRange; iter_rows with min/max bounds
    max_row = max(values_ws.max_row or 1, 1)
    max_col = max(values_ws.max_column or 1, 1)
    
    logger.info(f"Processing range: {max_row} rows x {max_col} columns")

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            address = a1_from_row_col(r, c)
            # Pull value from values_ws
            try:
                v = values_ws.cell(row=r, column=c).value
            except Exception:
                v = None
            # Detect formula
            formula = formulas_map.get(address)
            # Hyperlink
            hyperlink = hyperlinks_map.get(address)
            # Comment
            comment = comments_map.get(address)

            if formula:
                refs = regex_cross_sheet_refs(formula)
                for nm in refs:
                    if nm not in links_to:
                        links_to.append(nm)

            # Only include non-empty cells or those with metadata
            include = (v is not None) or (formula is not None) or (hyperlink is not None) or (comment is not None)
            if include:
                cells.append({
                    "address": address,
                    "row": r,
                    "col": c,
                    "value": v,
                    "formula": formula,
                    "data_type": None,
                    "hyperlink": hyperlink,
                    "comment": comment,
                })

    elapsed = time.time() - start_time
    logger.info(f"Extracted {len(cells)} cells, {len(links_to)} cross-sheet links in {elapsed:.2f}s")
    return cells, links_to


# -----------------------------
# Media & DrawingML extraction from .xlsx zip
# -----------------------------

def extract_media_and_drawings(xlsx_path: Path, out_media: Path, out_drawings: Path) -> Tuple[List[Dict], List[str]]:
    import zipfile

    logger.info("Extracting media and drawings from .xlsx")
    start_time = time.time()
    
    ensure_dir(out_media)
    ensure_dir(out_drawings)

    images: List[Dict] = []
    drawing_files: List[str] = []

    with zipfile.ZipFile(xlsx_path, 'r') as zf:
        for member in zf.namelist():
            if member.startswith('xl/media/') and not member.endswith('/'):
                target = out_media / Path(member).name
                with zf.open(member) as src, target.open('wb') as dst:
                    shutil.copyfileobj(src, dst)
                images.append({"path": str(target)})
            elif member.startswith('xl/drawings/') and member.endswith('.xml'):
                target = out_drawings / Path(member).name
                with zf.open(member) as src, target.open('wb') as dst:
                    shutil.copyfileobj(src, dst)
                drawing_files.append(member)

    elapsed = time.time() - start_time
    logger.info(f"Extracted {len(images)} media files, {len(drawing_files)} drawings in {elapsed:.2f}s")
    return images, drawing_files


def map_sheet_to_drawings(xlsx_path: Path) -> Dict[str, List[str]]:
    # Best-effort mapping using rels: xl/worksheets/_rels/sheetX.xml.rels → drawing rels
    import zipfile
    from lxml import etree

    mapping: Dict[str, List[str]] = {}

    with zipfile.ZipFile(xlsx_path, 'r') as zf:
        # Build sheetId → sheetName mapping
        workbook_xml = zf.read('xl/workbook.xml')
        wb_root = etree.fromstring(workbook_xml)
        ns = {
            'a': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
            'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
        }
        sheet_id_to_name: Dict[str, str] = {}
        for sheet in wb_root.xpath('.//a:sheets/a:sheet', namespaces=ns):
            name = sheet.get('name')
            r_id = sheet.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
            sheet_id_to_name[r_id] = name

        # Resolve workbook rels to sheet xml files
        wb_rels_xml = zf.read('xl/_rels/workbook.xml.rels')
        wb_rels_root = etree.fromstring(wb_rels_xml)
        rels_ns = {'r': 'http://schemas.openxmlformats.org/package/2006/relationships'}
        r_id_to_target: Dict[str, str] = {}
        for rel in wb_rels_root.xpath('.//r:Relationship', namespaces=rels_ns):
            r_id_to_target[rel.get('Id')] = rel.get('Target')  # e.g., worksheets/sheet1.xml

        # For each sheet rels, find drawing targets
        for r_id, target in r_id_to_target.items():
            if not target.startswith('worksheets/'):
                continue
            sheet_name = sheet_id_to_name.get(r_id)
            if not sheet_name:
                continue
            rels_path = f"xl/worksheets/_rels/{Path(target).name}.rels"
            if rels_path not in zf.namelist():
                continue
            rels_xml = zf.read(rels_path)
            rels_root = etree.fromstring(rels_xml)
            drawings_for_sheet: List[str] = []
            for rel in rels_root.xpath('.//r:Relationship', namespaces=rels_ns):
                target_rel = rel.get('Target')  # e.g., ../drawings/drawing1.xml
                if 'drawings/' in target_rel:
                    drawings_for_sheet.append(Path(target_rel).name)
            if drawings_for_sheet:
                mapping[sheet_name] = drawings_for_sheet

    return mapping


def get_ordered_sheet_info(xlsx_path: Path) -> Tuple[List[Dict], Dict[str, str]]:
    # Returns ordered list of dicts: {name, state, r_id, sheet_xml, sheet_rels}
    # and mapping r_id -> target (worksheets/sheetN.xml)
    import zipfile
    from lxml import etree

    with zipfile.ZipFile(xlsx_path, 'r') as zf:
        wb_root = etree.fromstring(zf.read('xl/workbook.xml'))
        ns = {
            'a': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
            'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
        }
        wb_rels_root = etree.fromstring(zf.read('xl/_rels/workbook.xml.rels'))
        rels_ns = {'r': 'http://schemas.openxmlformats.org/package/2006/relationships'}
        r_id_to_target: Dict[str, str] = {}
        for rel in wb_rels_root.xpath('.//r:Relationship', namespaces=rels_ns):
            r_id_to_target[rel.get('Id')] = rel.get('Target')

        ordered: List[Dict] = []
        for sheet in wb_root.xpath('.//a:sheets/a:sheet', namespaces=ns):
            name = sheet.get('name')
            r_id = sheet.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
            state = sheet.get('state', 'visible')
            target = r_id_to_target.get(r_id, '')  # e.g., worksheets/sheet1.xml
            sheet_xml = f"xl/{target}" if target else ''
            sheet_rels = f"xl/worksheets/_rels/{Path(target).name}.rels" if target else ''
            ordered.append({
                'name': name,
                'state': state,
                'r_id': r_id,
                'sheet_xml': sheet_xml,
                'sheet_rels': sheet_rels,
            })
    return ordered, r_id_to_target


def parse_sheet_meta_maps(xlsx_path: Path, sheet_xml_path: str, sheet_rels_path: str) -> Tuple[Dict[str, str], Dict[str, str], Dict[str, str]]:
    # Returns maps: formulas_map[address] = '=...', hyperlinks_map[address] = url, comments_map[address] = text
    import zipfile
    from lxml import etree
    from openpyxl.utils.cell import range_boundaries
    from openpyxl.utils import get_column_letter

    formulas: Dict[str, str] = {}
    hyperlinks: Dict[str, str] = {}
    comments: Dict[str, str] = {}

    with zipfile.ZipFile(xlsx_path, 'r') as zf:
        # Parse relationships to resolve hyperlinks and comments
        id_to_link: Dict[str, str] = {}
        comments_target: Optional[str] = None
        try:
            if sheet_rels_path and sheet_rels_path in zf.namelist():
                rels_root = etree.fromstring(zf.read(sheet_rels_path))
                rels_ns = {'r': 'http://schemas.openxmlformats.org/package/2006/relationships'}
                for rel in rels_root.xpath('.//r:Relationship', namespaces=rels_ns):
                    typ = rel.get('Type')
                    if typ and typ.endswith('/hyperlink'):
                        id_to_link[rel.get('Id')] = rel.get('Target')
                    if typ and typ.endswith('/comments'):
                        comments_target = rel.get('Target')  # e.g., ../comments1.xml
        except Exception:
            pass

        # Normalize comments target to absolute zip path
        if comments_target:
            target_path = comments_target
            if target_path.startswith('../'):
                target_path = 'xl/' + target_path.replace('../', '')
            elif not target_path.startswith('xl/'):
                target_path = 'xl/' + target_path
            try:
                if target_path in zf.namelist():
                    com_root = etree.fromstring(zf.read(target_path))
                    ns_c = {'c': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                    # Some comments parts may not use the namespace prefix; use local-names
                    for cmt in com_root.xpath('.//comment') + com_root.xpath('.//c:comment', namespaces=ns_c):
                        ref = cmt.get('ref')
                        if not ref:
                            continue
                        # Concatenate text nodes under comment/text
                        texts = []
                        for t in cmt.xpath('.//text//t') + cmt.xpath('.//c:text//c:t', namespaces=ns_c):
                            if t.text:
                                texts.append(t.text)
                        if texts:
                            comments[ref] = ''.join(texts)
            except Exception:
                pass

        # Parse sheet XML for formulas and hyperlink refs
        try:
            if sheet_xml_path and sheet_xml_path in zf.namelist():
                sh_root = etree.fromstring(zf.read(sheet_xml_path))
                ns = {'a': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                # Formulas: cell c with f child
                for c in sh_root.xpath('.//a:sheetData/a:row/a:c[a:f]', namespaces=ns):
                    addr = c.get('r')
                    f_el = c.find('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}f')
                    if addr and f_el is not None and f_el.text:
                        formulas[addr] = '=' + f_el.text
                # Hyperlinks section
                for h in sh_root.xpath('.//a:hyperlinks/a:hyperlink', namespaces=ns):
                    ref = h.get('ref')
                    r_id = h.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
                    location = h.get('location')
                    target = None
                    if r_id and r_id in id_to_link:
                        target = id_to_link[r_id]
                    elif location:
                        target = f"#{location}"
                    if ref and target:
                        # Expand ranges like A1:B3 to all cells
                        try:
                            min_col, min_row, max_col, max_row = range_boundaries(ref)
                            for rr in range(min_row, max_row + 1):
                                for cc in range(min_col, max_col + 1):
                                    addr = f"{get_column_letter(cc)}{rr}"
                                    hyperlinks[addr] = target
                        except Exception:
                            # Single cell like A1
                            hyperlinks[ref] = target
        except Exception:
            pass

    return formulas, hyperlinks, comments


# -----------------------------
# Visual export (xlwings/Excel for Mac)
# -----------------------------

def compute_print_area_including_shapes(sht) -> Tuple[int, int, int, int]:
    # Compute using cross-platform xlwings Range API (works on macOS)
    try:
        used = sht.used_range  # xlwings Range
        r1 = int(used.row)
        c1 = int(used.column)
        r2 = int(used.last_cell.row)
        c2 = int(used.last_cell.column)
    except Exception:
        # Fallback to a conservative area
        r1, c1, r2, c2 = 1, 1, 100, 50

    shapes_bounds = None
    try:
        shapes = sht.api.Shapes
        count = int(shapes.Count)
    except Exception:
        count = 0

    # Expand bounds by shapes' cell envelopes
    if count > 0:
        for i in range(1, count + 1):
            try:
                shp = shapes.Item(i)
                tl = shp.TopLeftCell
                br = shp.BottomRightCell
                sr1 = int(tl.Row)
                sc1 = int(tl.Column)
                sr2 = int(br.Row)
                sc2 = int(br.Column)
                if shapes_bounds is None:
                    shapes_bounds = [sr1, sc1, sr2, sc2]
                else:
                    shapes_bounds[0] = min(shapes_bounds[0], sr1)
                    shapes_bounds[1] = min(shapes_bounds[1], sc1)
                    shapes_bounds[2] = max(shapes_bounds[2], sr2)
                    shapes_bounds[3] = max(shapes_bounds[3], sc2)
            except Exception:
                # Ignore shapes without cell anchors
                continue

    # If used_range is minimal and shapes exist, rely on shapes bounds
    if (r1 == 1 and c1 == 1 and r2 <= 1 and c2 <= 1) and shapes_bounds is not None:
        r1, c1, r2, c2 = shapes_bounds

    # Safety padding to avoid edge cut-offs
    r2 = r2 + 2
    c2 = c2 + 2
    return r1, c1, r2, c2


def set_single_page_pagesetup(sht, r1: int, c1: int, r2: int, c2: int) -> None:
    from openpyxl.utils import get_column_letter

    # Build A1 range for PrintArea
    a1 = f"${get_column_letter(c1)}${r1}:${get_column_letter(c2)}${r2}"
    try:
        # macOS-safe page setup via xlwings wrapper
        sht.page_setup.print_area = a1
        sht.page_setup.fit_to_pages_wide = 1
        sht.page_setup.fit_to_pages_tall = 1
        logger.info(f"Set page setup: print_area={a1}, fit_to_pages 1x1")
    except Exception as e:
        logger.warning(f"Page setup via wrapper failed: {e}")


def export_sheet_visuals_python_fallback(values_ws, sheet_name: str, pdf_out: Path, png_out: Path) -> bool:
    """Export sheet as whole-sheet PDF/PNG using Python rendering (openpyxl + matplotlib)"""
    logger.info(f"Exporting visuals via Python rendering for sheet: {sheet_name}")
    start_time = time.time()
    
    try:
        import matplotlib.pyplot as plt
        import matplotlib.patches as patches
        from matplotlib.backends.backend_pdf import PdfPages
        import numpy as np
        
        # Get sheet dimensions
        max_row = max(values_ws.max_row or 1, 1)
        max_col = max(values_ws.max_column or 1, 1)
        
        # Create figure with appropriate size (A4-like proportions)
        fig_width = max(8.5, max_col * 0.8)  # ~0.8 inches per column
        fig_height = max(11, max_row * 0.2)  # ~0.2 inches per row
        fig, ax = plt.subplots(figsize=(fig_width, fig_height))
        
        # Set up the plot area
        ax.set_xlim(0, max_col)
        ax.set_ylim(0, max_row)
        ax.set_aspect('equal')
        ax.invert_yaxis()  # Excel-style: row 1 at top
        
        # Draw grid
        for i in range(max_col + 1):
            ax.axvline(x=i, color='lightgray', linewidth=0.5)
        for i in range(max_row + 1):
            ax.axhline(y=i, color='lightgray', linewidth=0.5)
        
        # Fill cells with data
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                try:
                    cell = values_ws.cell(row=row, column=col)
                    if cell.value is not None:
                        # Draw cell background
                        rect = patches.Rectangle(
                            (col-1, row-1), 1, 1,
                            linewidth=0.5, edgecolor='black',
                            facecolor='white', alpha=0.8
                        )
                        ax.add_patch(rect)
                        
                        # Add text (truncate if too long)
                        text = str(cell.value)[:20] + "..." if len(str(cell.value)) > 20 else str(cell.value)
                        ax.text(col-0.5, row-0.5, text, 
                               ha='center', va='center', fontsize=8, wrap=True)
                except Exception:
                    continue
        
        # Remove axes
        ax.set_xticks([])
        ax.set_yticks([])
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.spines['bottom'].set_visible(False)
        ax.spines['left'].set_visible(False)
        
        # Save as PDF (whole sheet)
        logger.info(f"Exporting PDF to {pdf_out.name}")
        with PdfPages(pdf_out) as pdf:
            pdf.savefig(fig, bbox_inches='tight', pad_inches=0.1)
        
        # Save as PNG (whole sheet)
        logger.info(f"Exporting PNG to {png_out.name}")
        fig.savefig(png_out, dpi=300, bbox_inches='tight', pad_inches=0.1)
        
        plt.close(fig)
        
        elapsed = time.time() - start_time
        logger.info(f"Python visual export completed in {elapsed:.2f}s")
        return True
        
    except Exception as e:
        logger.warning(f"Python visual export failed: {e}")
        return False


def _applescript_get_excel_window_bounds() -> Optional[Tuple[int, int, int, int]]:
    """Return (x, y, w, h) of the front Excel window via AppleScript/System Events."""
    script = (
        'tell application "System Events"\n'
        '  tell application process "Microsoft Excel"\n'
        '    if (count of windows) is 0 then return ""\n'
        '    set theWindow to front window\n'
        '    set {x, y} to position of theWindow\n'
        '    set {w, h} to size of theWindow\n'
        '    return (x as string) & "," & (y as string) & "," & (w as string) & "," & (h as string)\n'
        '  end tell\n'
        'end tell'
    )
    try:
        result = subprocess.run(["osascript", "-e", script], capture_output=True, text=True, timeout=5)
        if result.returncode != 0:
            return None
        out = result.stdout.strip()
        if not out:
            return None
        parts = out.split(",")
        if len(parts) != 4:
            return None
        x, y, w, h = map(int, parts)
        return x, y, w, h
    except Exception:
        return None


def _capture_active_excel_window_png(png_out: Path) -> bool:
    """Capture the active Excel window into a PNG using macOS screencapture."""
    # Prefer capturing by window id to avoid coordinate scaling issues
    win_id = _applescript_get_excel_window_id()
    if win_id is not None:
        try:
            cmd = ["screencapture", "-x", "-l", str(win_id), str(png_out)]
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=10)
            if result.returncode == 0 and png_out.exists():
                return True
        except Exception:
            pass
    # Fallback to rectangle bounds
    bounds = _applescript_get_excel_window_bounds()
    if bounds:
        x, y, w, h = bounds
        try:
            cmd = ["screencapture", "-x", f"-R{x},{y},{w},{h}", str(png_out)]
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=10)
            return result.returncode == 0 and png_out.exists()
        except Exception:
            return False
    return False


def _applescript_get_excel_window_id() -> Optional[int]:
    """Return CGWindowID-like id of the front Excel window via AppleScript/System Events."""
    script = (
        'tell application "Microsoft Excel" to activate\n'
        'tell application "System Events"\n'
        '  tell application process "Microsoft Excel"\n'
        '    if (count of windows) is 0 then return ""\n'
        '    set theWindow to front window\n'
        '    try\n'
        '      return id of theWindow\n'
        '    on error\n'
        '      return ""\n'
        '    end try\n'
        '  end tell\n'
        'end tell'
    )
    try:
        result = subprocess.run(["osascript", "-e", script], capture_output=True, text=True, timeout=5)
        if result.returncode != 0:
            return None
        out = result.stdout.strip()
        if not out:
            return None
        return int(out)
    except Exception:
        return None


def _png_to_pdf(png_path: Path, pdf_path: Path, orientation: Optional[str] = None) -> bool:
    try:
        from PIL import Image
        img = Image.open(png_path)
        # Orientation inference if not provided
        if orientation is None:
            orientation = "landscape" if img.width >= img.height else "portrait"
        # Save as single-page PDF
        rgb = img.convert("RGB")
        rgb.save(pdf_path, "PDF", resolution=300)
        return True
    except Exception as e:
        logger.warning(f"PNG to PDF conversion failed: {e}")
        return False


def _copy_picture_menu_to_clipboard() -> bool:
    """Trigger Excel's 'Copy Picture…' via UI scripting (macOS) and confirm dialog."""
    # Handles both three-dot and ellipsis variants
    script = (
        'tell application "System Events"\n'
        '  tell application process "Microsoft Excel"\n'
        '    if not (exists menu bar 1) then return false\n'
        '    try\n'
        '      click menu item "Copy Picture…" of menu "Edit" of menu bar item "Edit" of menu bar 1\n'
        '    on error\n'
        '      try\n'
        '        click menu item "Copy Picture..." of menu "Edit" of menu bar item "Edit" of menu bar 1\n'
        '      on error\n'
        '        return false\n'
        '      end try\n'
        '    end try\n'
        '    delay 0.2\n'
        '    if exists sheet 1 of window 1 then\n'
        '      try\n'
        '        click button "OK" of sheet 1 of window 1\n'
        '      end try\n'
        '    end if\n'
        '    return true\n'
        '  end tell\n'
        'end tell'
    )
    try:
        result = subprocess.run(["osascript", "-e", script], capture_output=True, text=True, timeout=5)
        return result.returncode == 0 and result.stdout.strip() == 'true'
    except Exception:
        return False


def _save_clipboard_image_to_png(png_out: Path) -> bool:
    try:
        from PIL import ImageGrab
        grab = ImageGrab.grabclipboard()
        if grab is None:
            return False
        grab.save(png_out, format='PNG')
        return png_out.exists()
    except Exception:
        return False


def excel_zoom_fit_selection(sheet_name: str) -> bool:
    """Use Excel VBA via AppleScript to Zoom To Selection and FitToPages 1x1."""
    try:
        vba_zoom = 'On Error Resume Next: With ActiveSheet: .UsedRange.Select: Application.CommandBars.ExecuteMso("ZoomToSelection"): .PageSetup.FitToPagesWide=1: .PageSetup.FitToPagesTall=1: End With'
        osa = f'''
        tell application "Microsoft Excel"
          activate
          try
            tell active workbook
              tell sheet "{sheet_name}" to activate
            end tell
          on error
            return "ERR_NO_SHEET"
          end try
          try
            do Visual Basic "{vba_zoom}"
            return "OK"
          on error errMsg
            return "ERR:" & errMsg
          end try
        end tell
        '''
        res = subprocess.run(["osascript", "-e", osa], capture_output=True, text=True, timeout=6)
        return res.returncode == 0 and "OK" in (res.stdout or "")
    except Exception:
        return False


def _col_width_to_pixels(col_width: Optional[float]) -> int:
    # Approx conversion from Excel column width to pixels
    if col_width is None:
        col_width = 8.43
    try:
        return int(round(7 * col_width + 5))
    except Exception:
        return 64


def _row_height_to_pixels(row_height: Optional[float]) -> int:
    # Row height given in points; 1 point = 1/72 inch; assume 96 dpi → px = pt * 96/72
    if row_height is None:
        row_height = 15.0
    try:
        return int(round(row_height * 96.0 / 72.0))
    except Exception:
        return 20


def _emu_to_px(emu: Optional[int]) -> int:
    if not emu:
        return 0
    try:
        return int(round((emu / 914400.0) * 96.0))
    except Exception:
        return 0


def export_sheet_visuals_overlay_from_drawing(values_ws, sheet_name: str, xlsx_path: Path, pdf_out: Path, png_out: Path, area: Optional[Tuple[int, int, int, int]] = None) -> bool:
    """Reconstruct a visual by overlaying images from DrawingML anchors onto a canvas sized to the sheet area.
    This includes SmartArt/images/charts as rendered images, but not cell text.
    """
    logger.info(f"Exporting visuals via DrawingML overlay for sheet: {sheet_name}")
    try:
        import zipfile
        from lxml import etree
        from PIL import Image, ImageDraw
        from openpyxl.utils import get_column_letter

        # Determine sheet drawings
        mapping = map_sheet_to_drawings(xlsx_path)
        drawings = mapping.get(sheet_name, [])
        if not drawings:
            logger.warning("No drawing parts mapped to sheet; overlay cannot proceed")
            return False

        # Compute canvas size using area bounds or sheet dimensions
        if area is not None:
            r1, c1, r2, c2 = area
        else:
            r1, c1 = 1, 1
            r2 = max(values_ws.max_row or 1, 1)
            c2 = max(values_ws.max_column or 1, 1)

        # Sum pixel widths/heights across target area
        total_width_px = 0
        col_offsets_px: List[int] = [0]
        for c in range(c1, c2 + 1):
            key = get_column_letter(c)
            w_px = _col_width_to_pixels(getattr(values_ws.column_dimensions.get(key, None), 'width', None) if values_ws.column_dimensions and key in values_ws.column_dimensions else None)
            total_width_px += w_px
            col_offsets_px.append(total_width_px)

        total_height_px = 0
        row_offsets_px: List[int] = [0]
        for r in range(r1, r2 + 1):
            h_px = _row_height_to_pixels(getattr(values_ws.row_dimensions.get(r, None), 'height', None) if values_ws.row_dimensions and r in values_ws.row_dimensions else None)
            total_height_px += h_px
            row_offsets_px.append(total_height_px)

        # Create blank white canvas
        canvas = Image.new('RGB', (max(1, total_width_px), max(1, total_height_px)), (255, 255, 255))
        draw = ImageDraw.Draw(canvas)

        with zipfile.ZipFile(xlsx_path, 'r') as zf:
            for drawing_name in drawings:
                drawing_path = f"xl/drawings/{drawing_name}"
                rels_path = f"xl/drawings/_rels/{Path(drawing_name).name}.rels"
                if drawing_path not in zf.namelist():
                    continue
                rels_map: Dict[str, str] = {}
                if rels_path in zf.namelist():
                    rels_root = etree.fromstring(zf.read(rels_path))
                    rels_ns = {'r': 'http://schemas.openxmlformats.org/package/2006/relationships'}
                    for rel in rels_root.xpath('.//r:Relationship', namespaces=rels_ns):
                        if rel.get('Type', '').endswith('/image'):
                            rels_map[rel.get('Id')] = rel.get('Target').replace('../', 'xl/')

                root = etree.fromstring(zf.read(drawing_path))
                ns = {
                    'a': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
                    'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
                    'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
                }
                # twoCellAnchor holds from/to cells and embedded image via r:embed
                for anchor in root.xpath('.//xdr:twoCellAnchor', namespaces=ns):
                    pic = anchor.find('.//xdr:pic', namespaces=ns)
                    if pic is None:
                        continue
                    blip = pic.find('.//a:blip', namespaces={'a': 'http://schemas.openxmlformats.org/drawingml/2006/main'})
                    if blip is None:
                        continue
                    r_id = blip.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed')
                    media_path = rels_map.get(r_id)
                    if not media_path or media_path not in zf.namelist():
                        continue

                    # from cell
                    from_node = anchor.find('xdr:from', namespaces=ns)
                    to_node = anchor.find('xdr:to', namespaces=ns)
                    if from_node is None or to_node is None:
                        continue
                    try:
                        from_col = int(from_node.find('xdr:col', namespaces=ns).text) + 1
                        from_colOff = int(from_node.find('xdr:colOff', namespaces=ns).text or '0')
                        from_row = int(from_node.find('xdr:row', namespaces=ns).text) + 1
                        from_rowOff = int(from_node.find('xdr:rowOff', namespaces=ns).text or '0')

                        to_col = int(to_node.find('xdr:col', namespaces=ns).text) + 1
                        to_colOff = int(to_node.find('xdr:colOff', namespaces=ns).text or '0')
                        to_row = int(to_node.find('xdr:row', namespaces=ns).text) + 1
                        to_rowOff = int(to_node.find('xdr:rowOff', namespaces=ns).text or '0')
                    except Exception:
                        continue

                    # Map to pixel positions within the canvas area (r1..r2, c1..c2)
                    if from_col < c1 or to_col > c2 or from_row < r1 or to_row > r2:
                        # If image outside area, skip or clamp (here we skip)
                        continue

                    # Position from area origin
                    x0 = col_offsets_px[from_col - c1] + _emu_to_px(from_colOff)
                    y0 = row_offsets_px[from_row - r1] + _emu_to_px(from_rowOff)
                    x1 = col_offsets_px[to_col - c1] + _emu_to_px(to_colOff)
                    y1 = row_offsets_px[to_row - r1] + _emu_to_px(to_rowOff)

                    # Load and paste image scaled to box
                    try:
                        with zf.open(media_path) as imgf:
                            img = Image.open(imgf).convert('RGBA')
                            box_w = max(1, x1 - x0)
                            box_h = max(1, y1 - y0)
                            img_resized = img.resize((box_w, box_h))
                            canvas.paste(img_resized, (x0, y0), img_resized)
                    except Exception as e:
                        logger.warning(f"Failed to place media {media_path}: {e}")

        # Save outputs
        canvas.save(png_out, format='PNG')
        _png_to_pdf(png_out, pdf_out)
        logger.info("DrawingML overlay export completed (PNG first → PDF)")
        return True
    except Exception as e:
        logger.warning(f"DrawingML overlay export failed: {e}")
        return False


def export_sheet_visuals_mac(sht, pdf_out: Path, png_out: Path, values_ws=None) -> None:
    from openpyxl.utils import get_column_letter

    logger.info(f"Exporting visuals for sheet: {sht.name}")
    start_time = time.time()
    
    excel_success = False
    screenshot_success = False

    # 1) Preferred: Use Excel's Copy Picture… via UI scripting and clipboard capture
    try:
        from openpyxl.utils import get_column_letter
        logger.info("Attempting Copy Picture via UI and clipboard (preferred)")
        # Compute shape-aware print area and select it
        r1, c1, r2, c2 = compute_print_area_including_shapes(sht)
        a1 = f"${get_column_letter(c1)}${r1}:${get_column_letter(c2)}${r2}"
        sht.range(a1).select()
        _time.sleep(0.2)
        # Zoom to selection to fit on screen before copying
        excel_zoom_fit_selection(sht.name)
        _time.sleep(0.2)
        if _copy_picture_menu_to_clipboard():
            if _save_clipboard_image_to_png(png_out):
                # Convert to PDF
                _png_to_pdf(png_out, pdf_out)
                logger.info("Clipboard-based image export completed (PNG first → PDF)")
                screenshot_success = True
        else:
            logger.warning("Could not trigger Copy Picture via menu")
    except Exception as e:
        logger.warning(f"Copy Picture clipboard path failed: {e}")

    # 2) If Copy Picture fails, try AppleScript screenshot of active Excel window
    if not screenshot_success:
        try:
            from openpyxl.utils import get_column_letter
            # Ensure selection and zoom before screenshot too
            r1s, c1s, r2s, c2s = compute_print_area_including_shapes(sht)
            a1s = f"${get_column_letter(c1s)}${r1s}:${get_column_letter(c2s)}${r2s}"
            sht.range(a1s).select()
            _time.sleep(0.2)
            excel_zoom_fit_selection(sht.name)
            _time.sleep(0.3)

            sht.activate()
            _time.sleep(0.3)
            logger.info("Capturing sheet screenshot via AppleScript")
            screenshot_success = _capture_active_excel_window_png(png_out)
            if screenshot_success:
                # Detect orientation from sheet if possible
                try:
                    orient = None
                    ps = getattr(sht, 'page_setup', None)
                    if ps is not None:
                        o = getattr(ps, 'orientation', None)
                        if o == 2:
                            orient = 'landscape'
                        elif o == 1:
                            orient = 'portrait'
                    _png_to_pdf(png_out, pdf_out, orient)
                    logger.info("Screenshot export completed (PNG first → PDF)")
                except Exception:
                    pass
        except Exception as e:
            logger.warning(f"AppleScript screenshot failed: {e}")

    # Try Excel automation (may fail on macOS)
    try:
        # PASS 1
        logger.info("Computing shape-aware print area")
        r1, c1, r2, c2 = compute_print_area_including_shapes(sht)
        logger.info(f"Print area: {r1},{c1} to {r2},{c2}")
        
        set_single_page_pagesetup(sht, r1, c1, r2, c2)

        # Export PDF with fallback
        try:
            logger.info(f"Exporting PDF to {pdf_out.name}")
            sht.api.ExportAsFixedFormat(0, str(pdf_out))  # 0 == xlTypePDF
            excel_success = True
        except Exception as e:
            logger.warning(f"Primary PDF export failed: {e}; trying ActiveSheet fallback")
            try:
                sht.book.api.ActiveSheet.ExportAsFixedFormat(0, str(pdf_out))
                excel_success = True
            except Exception as e2:
                logger.warning(f"Fallback PDF export failed: {e2}")

        # Export PNG via CopyPicture + Chart.Export
        if excel_success:
            logger.info(f"Exporting PNG to {png_out.name}")
            rng = sht.range(f"${get_column_letter(c1)}${r1}:${get_column_letter(c2)}${r2}")
            rng.api.CopyPicture(Appearance=1, Format=2)  # xlScreen, xlPicture

            def try_export_png(width: int, height: int) -> bool:
                try:
                    ch = sht.api.ChartObjects().Add(10, 10, width, height)
                    ch.Activate()
                    sht.api.Paste()
                    ch.Chart.Export(str(png_out))
                    ch.Delete()
                    return True
                except Exception as e:
                    logger.warning(f"PNG export attempt {width}x{height} failed: {e}")
                    try:
                        ch.Delete()
                    except Exception:
                        pass
                    return False

            if not try_export_png(8000, 4500):
                # Retry with alternative size
                try_export_png(6000, 4000)
                
    except Exception as e:
        logger.warning(f"Excel automation failed: {e}")
    
    # If UI-based methods failed, try DrawingML overlay based on anchors and media
    if not screenshot_success or not pdf_out.exists() or not png_out.exists():
        try:
            logger.info("Attempting DrawingML overlay export")
            # Compute area again for sizing
            r1o, c1o, r2o, c2o = compute_print_area_including_shapes(sht)
            xlsx_path = Path(sht.book.fullname)
            if values_ws is not None:
                overlay_ok = export_sheet_visuals_overlay_from_drawing(values_ws, sht.name, xlsx_path, pdf_out, png_out, (r1o, c1o, r2o, c2o))
            else:
                overlay_ok = False
            screenshot_success = screenshot_success or overlay_ok
        except Exception as e:
            logger.warning(f"DrawingML overlay path failed: {e}")

    # Final fallback: Python rendering of cells only
    if not screenshot_success or not pdf_out.exists() or not png_out.exists():
        logger.info("Falling back to Python cell rendering for visual export")
        try:
            if values_ws is not None:
                export_sheet_visuals_python_fallback(values_ws, sht.name, pdf_out, png_out)
            else:
                logger.warning("No values worksheet available for Python fallback")
        except Exception as e:
            logger.warning(f"Python rendering fallback failed: {e}")
    
    elapsed = time.time() - start_time
    logger.info(f"Visual export completed in {elapsed:.2f}s")


# -----------------------------
# OCR (optional)
# -----------------------------

def run_ocr_if_enabled(png_path: Path, ocr_out: Path) -> None:
    logger.info(f"Running OCR on {png_path.name}")
    start_time = time.time()
    
    try:
        import pytesseract
        from PIL import Image
    except Exception:
        logger.warning("OCR dependencies not available")
        return

    try:
        text = pytesseract.image_to_string(Image.open(png_path))
        ocr_out.write_text(text, encoding="utf-8")
        elapsed = time.time() - start_time
        logger.info(f"OCR completed in {elapsed:.2f}s ({len(text)} characters)")
    except Exception as e:
        logger.warning(f"OCR failed: {e}")
        # Best-effort OCR only; ignore failures
        pass


# -----------------------------
# Main pipeline
# -----------------------------

def preprocess_workbook(xlsx_path: Path, out_dir: Path, num_sheets: int, enable_ocr: bool) -> None:
    logger.info("=" * 60)
    logger.info("EXCEL PREPROCESSING PIPELINE STARTED")
    logger.info("=" * 60)
    logger.info(f"Workbook: {xlsx_path}")
    logger.info(f"Output: {out_dir}")
    logger.info(f"Processing first {num_sheets} sheets")
    logger.info(f"OCR enabled: {enable_ocr}")
    
    pipeline_start = time.time()
    
    ensure_dir(out_dir)

    # Copy original workbook
    logger.info("Copying original workbook")
    copied_xlsx = out_dir / xlsx_path.name
    if not copied_xlsx.exists():
        shutil.copy2(xlsx_path, copied_xlsx)

    # Subdirs
    sheets_dir = out_dir / "sheets"
    assets_pdf_dir = out_dir / "assets" / "pdf"
    assets_png_dir = out_dir / "assets" / "png"
    assets_ocr_dir = out_dir / "assets" / "ocr"
    extracted_media_dir = out_dir / "extracted" / "media"
    extracted_drawings_dir = out_dir / "extracted" / "drawings"

    for d in [sheets_dir, assets_pdf_dir, assets_png_dir, extracted_media_dir, extracted_drawings_dir]:
        ensure_dir(d)
    if enable_ocr:
        ensure_dir(assets_ocr_dir)

    workbook_id = compute_sha256(xlsx_path)

    # Load workbook for values only (fast)
    from openpyxl import __version__ as openpyxl_version  # noqa: F401
    values_wb = load_values_workbook(xlsx_path)

    # Extract media/drawings (global)
    images_global, drawing_files_global = extract_media_and_drawings(xlsx_path, extracted_media_dir, extracted_drawings_dir)
    sheet_to_drawings = {}
    try:
        logger.info("Mapping sheets to drawings")
        sheet_to_drawings = map_sheet_to_drawings(xlsx_path)
        logger.info(f"Mapped {len(sheet_to_drawings)} sheets to drawings")
    except Exception as e:
        logger.warning(f"Could not map sheets to drawings: {e}")
        sheet_to_drawings = {}

    # Manifest structure
    manifest = {
        "workbook": str(xlsx_path),
        "workbook_id": workbook_id,
        "created_at": int(time.time()),
        "sheets": [],
        "extracted": {
            "media_dir": str(extracted_media_dir),
            "drawings_dir": str(extracted_drawings_dir),
        },
    }

    # Iterate sheets (first N visible worksheets)
    processed = 0
    ordered_info, _ = get_ordered_sheet_info(xlsx_path)

    # Visual export via Excel only if available
    can_visual = True
    try:
        import xlwings as xw  # noqa: F401
        logger.info("xlwings available - visual export enabled")
    except Exception:
        logger.warning("xlwings not available - visual export disabled")
        can_visual = False

    app = None
    wb_xlw = None
    if can_visual:
        logger.info("Opening Excel for visual export")
        import xlwings as xw
        app = xw.App(visible=False, add_book=False)
        try:
            wb_xlw = app.books.open(str(xlsx_path))
            logger.info("Excel workbook opened successfully")
        except Exception as e:
            logger.warning(f"Could not open Excel workbook: {e}")
            wb_xlw = None
            can_visual = False

    try:
        logger.info(f"Processing {len(ordered_info)} available sheets")
        for idx, info in enumerate(ordered_info):
            if processed >= num_sheets:
                break
            ws_name = info['name']
            state = info['state']
            sheet_xml = info['sheet_xml']
            sheet_rels = info['sheet_rels']

            logger.info(f"Processing sheet {processed + 1}/{num_sheets}: {ws_name}")

            if not is_effectively_visible_state(state):
                logger.info(f"Skipping hidden sheet: {ws_name}")
                continue

            # Structured extraction
            try:
                values_ws = values_wb[ws_name]
            except KeyError:
                logger.warning(f"Sheet not found in values workbook: {ws_name}")
                continue

            formulas_map, hyperlinks_map, comments_map = parse_sheet_meta_maps(xlsx_path, sheet_xml, sheet_rels)
            cells, links_to = extract_sheet_structured(values_ws, ws_name, formulas_map, hyperlinks_map, comments_map)
            named_ranges_all = parse_defined_names_from_workbook(xlsx_path)

            # Drawing files mapped to this sheet
            smartart_list = sheet_to_drawings.get(ws_name, [])

            # Visual export
            pdf_path = (assets_pdf_dir / f"{sanitize_name(ws_name)}.pdf") if can_visual else None
            png_path = (assets_png_dir / f"{sanitize_name(ws_name)}.png") if can_visual else None
            ocr_path = (assets_ocr_dir / f"{sanitize_name(ws_name)}.txt") if (enable_ocr and can_visual) else None

            if can_visual and wb_xlw is not None:
                try:
                    sht = wb_xlw.sheets[ws_name]
                    export_sheet_visuals_mac(sht, pdf_path, png_path, values_ws)
                    if enable_ocr and ocr_path is not None and png_path is not None and png_path.exists():
                        run_ocr_if_enabled(png_path, ocr_path)
                except Exception as e:
                    # If visual export fails, continue with structured only
                    sys.stderr.write(f"[warn] Visual export failed for sheet '{ws_name}': {e}\n")
                    pdf_path = None
                    png_path = None
                    ocr_path = None

            sheet_obj = ExtractedSheet(
                workbook_id=workbook_id,
                sheet_index=idx,
                sheet_name=ws_name,
                cells=cells,
                named_ranges=named_ranges_all,
                images=images_global,
                smartart_drawings=smartart_list,
                assets={
                    "pdf": str(pdf_path) if pdf_path else None,
                    "png": str(png_path) if png_path else None,
                    "ocr": str(ocr_path) if (enable_ocr and ocr_path and ocr_path.exists()) else None,
                },
                links_to=links_to,
            )

            # Write per-sheet JSON
            logger.info(f"Writing JSON for sheet: {ws_name}")
            sheet_json_path = sheets_dir / f"{sanitize_name(ws_name)}.json"
            with sheet_json_path.open("w", encoding="utf-8") as f:
                json.dump(sheet_obj.__dict__, f, ensure_ascii=False, indent=2)

            # Append to manifest
            manifest["sheets"].append({
                "sheet_index": idx,
                "sheet_name": ws_name,
                "json": str(sheet_json_path),
                "pdf": str(pdf_path) if pdf_path else None,
                "png": str(png_path) if png_path else None,
                "ocr": str(ocr_path) if (enable_ocr and ocr_path and ocr_path.exists()) else None,
                "smartart_drawings": smartart_list,
            })

            processed += 1
            logger.info(f"Completed sheet {processed}/{num_sheets}: {ws_name}")

    finally:
        try:
            values_wb.close()
        except Exception:
            pass
        if wb_xlw is not None:
            try:
                wb_xlw.close()
            except Exception:
                pass
        if 'app' in locals() and app is not None:
            try:
                app.quit()
            except Exception:
                pass

    # Write manifest
    logger.info("Writing manifest.json")
    with (out_dir / "manifest.json").open("w", encoding="utf-8") as f:
        json.dump(manifest, f, ensure_ascii=False, indent=2)
    
    pipeline_elapsed = time.time() - pipeline_start
    logger.info("=" * 60)
    logger.info("PIPELINE COMPLETED SUCCESSFULLY")
    logger.info("=" * 60)
    logger.info(f"Total runtime: {pipeline_elapsed:.2f}s")
    logger.info(f"Processed {processed} sheets")
    logger.info(f"Output directory: {out_dir}")
    logger.info("=" * 60)


# -----------------------------
# CLI
# -----------------------------

def parse_args(argv: Optional[List[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Excel Preprocess (macOS): structured + visual export for first N sheets")
    parser.add_argument("--xlsx", type=str, default="", help="Path to workbook (.xlsx)")
    parser.add_argument("--out", type=str, default="preprocessed_output", help="Output directory")
    parser.add_argument("--sheets", type=int, default=5, help="Number of sheets to process")
    parser.add_argument("--ocr", type=int, default=0, help="Enable OCR on PNGs (0/1)")
    return parser.parse_args(argv)


def main(argv: Optional[List[str]] = None) -> int:
    args = parse_args(argv)

    # Resolve workbook path defaulting logic
    root = Path.cwd()
    preferred = root / "first5sheet.xlsx"
    fallback = root / "first5sheets.xlsx"
    xlsx_path = Path(args.xlsx).expanduser() if args.xlsx else (preferred if preferred.exists() else fallback)

    if not xlsx_path.exists():
        print(f"ERROR: Workbook not found: {xlsx_path}")
        return 2

    out_dir = Path(args.out).expanduser()
    num_sheets = max(1, int(args.sheets))
    enable_ocr = bool(int(args.ocr))

    preprocess_workbook(xlsx_path, out_dir, num_sheets, enable_ocr)
    print(f"✓ Preprocess complete. Output at: {out_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
